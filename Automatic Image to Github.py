#Please make a copy of the folder I have snt over and use that, as the folder will become clogged with all images you extract.
#The excel must have Property names in one column, and pasted images in another column. Which columns don't matter, you can adjust the code below to match.
#For an image to correspond with a property name, the top left corner of the image must be in the same row as the property name.
#I will work on a different approach for when images are not in this format, but for now it works.
#Send a grad a feeder (or any sheet) and just ask them to paste the images into the image column for their corresponding property names.
#Then run this excel through the code.
#If the images are formatted as "place in cell", this will not work. Again, I will try to build something less tempramental.
#This did 175 images to github links in around 5-7 minutes, so it is very quick compared to manual. Give me a message if things break.

import pandas as pd
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import os

#You must run the command "pip install openpyxl-image-loader pandas openpyxl" in your terminal to install the required libraries before running this code.

# Configuration variables
excel_path = r"O:\NIA\PORTFOLIOS TEAM\Portfolios - 2026\CT UK PAIF\4. PBI\CTI Asset Book 12_05_26.xlsx" #Paste the file path to the excel here, the r at the start is important.
sheet_name = 'Sheet1'
image_column = 'K' # Column containing the photos
id_column = 'A'    # Column containing property names or IDs to name the photo files
github_user = 'H1470' #Add your github user here
github_repo = 'Automatic-Image-To-Github---CTI-Final' #Add the repository name here
branch = 'main' #Leave this alone probably
folder = 'Images'  # This is the Images folder within this Folder, it will be populated with the image files
filename_prefix = 'Imtwo'  # Add a prefix to the start of each filename (e.g., 'CTI_', 'Property_'), leave empty for no prefix

os.makedirs(folder, exist_ok=True)

# Load workbook and image loader
wb = openpyxl.load_workbook(excel_path)
sheet = wb[sheet_name]
image_loader = SheetImageLoader(sheet) # Maps images to their specific cells

output_data = []

# Iterate through the rows to extract and map
print(f"Starting to iterate from row 2 to {sheet.max_row}...")
images_found = 0
rows_with_id = 0

for row in range(2, sheet.max_row + 1):
    id_val = str(sheet[f"{id_column}{row}"].value)
    img_cell = f"{image_column}{row}"
    
    if id_val != "None":
        rows_with_id += 1
    
    if id_val != "None" and image_loader.image_in(img_cell):
        images_found += 1
        # Extract and save image locally
        image = image_loader.get(img_cell)
        # Sanitize filename: remove invalid Windows characters and whitespace
        invalid_chars = '<>:"/\\|?*\n\r\t'
        safe_id = id_val.strip()  # Remove leading/trailing whitespace
        for char in invalid_chars:
            safe_id = safe_id.replace(char, '_')
        filename = f"{filename_prefix}{safe_id}.png"
        image.save(f"{folder}/{filename}")
        
        # Construct the raw GitHub URL predictably
        raw_url = f"https://raw.githubusercontent.com/{github_user}/{github_repo}/{branch}/{folder}/{filename}"
        output_data.append({"Property ID": id_val, "Raw GitHub Link": raw_url})

print(f"\nDebug info:")
print(f"  Rows with ID in column B: {rows_with_id}")
print(f"  Images found and processed: {images_found}")

# Export the ready-to-paste URLs
pd.DataFrame(output_data).to_excel("Imtwo_github_image_links.xlsx", index=False)
print("Extraction and link generation complete!")

#Once the code has run, there will be a message in the terminal saying it has completed. Then head to the github section on the left hand br, and click commit.
#Then press sync changes and wait for it all to load.
#An excel file will now be in the foler, containing property names and their corresponding github raw links. The Images will be on github and the links will already work.
#Paste them in to your feeder excel and Voila

#NOTES

#Large images will not break the program, but their links won't work. Compress images that are too large to avoid this
#This is a WIP, I will build more around it to make sure the process is as clean as possible.